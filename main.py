
from openpyxl import load_workbook
import lxml
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By  # 指定 HTML 文件中 DOM 标签元素
from selenium.webdriver.support.ui import WebDriverWait  # 等待网页加载完成
from selenium.webdriver.support import expected_conditions as EC  # 指定等待网页加载结束条件
from selenium.webdriver.common.keys import Keys
import pyperclip
import time


def find_coordinate(station):
    brs.get('http://api.map.baidu.com/lbsapi/getpoint/index.html')
    WebDriverWait(brs, 15).until(EC.visibility_of_element_located((By.ID, "localvalue")))
    text = brs.find_element_by_id('localvalue')
    text.clear()
    text.send_keys(station)
    text.send_keys(Keys.RETURN)  # 输入站名搜索
    WebDriverWait(brs, 5).until(EC.visibility_of_element_located((By.ID, "no0")))
    # time.sleep(0.1)  # 初步测试这是满足同城和异城切换的最短响应时间
    try:
        target = brs.find_element_by_xpath("//*[contains(text(),'途径地铁')]")
        target.click()  # 定位地铁站元素
        time.sleep(0.1)
        copy = brs.find_element_by_id('copyButton')
        copy.click()  # 复制坐标
    except:  # 可以使得在发生异常时不影响后续元素的查找，集中处理异常元素
        pyperclip.copy('-1, -1')


# webdriver初始化
brs = webdriver.Chrome()

# 电子表初始化
wb = load_workbook('subway.xlsx')
ws = wb["Sheet1"]

for i in range(107, 513):
    key = ws['D' + str(i)]
    city = ws['B' + str(i)].value[:-1] + ' '
    find_coordinate(city + key.value + '站')
    coordinate = pyperclip.paste()  # 从剪切板获取刚复制的坐标
    cod = coordinate.split(',')
    ws['E' + str(i)] = float(cod[0])
    ws['F' + str(i)] = float(cod[1])
    wb.save('subway.xlsx')

brs.close()
wb.save('subway.xlsx')
